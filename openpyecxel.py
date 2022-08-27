import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

wb=load_workbook("salary.xlsx")

ws=wb['Sheet1']

ws['A1']="Salary Components"
ws['A1'].font=Font(bold=True)

ws['B1']="Yealy"
ws['B1'].font=Font(bold=True)

ws['C1']="Monthly"
ws['C1'].font=Font(bold=True)

ws['A2']="Base Salary"
ws['A3']="HRA"
ws['A4']="SA"
ws['A5']="PF"

ws['A6']="Basic Salary"
ws['A6'].font=Font(bold=True)

ws['A7']="Yearly Bonus"
ws['A8']="Other Allowance"

ws['A9']="CTC"
ws['A9'].font=Font(bold=True)



ws['B6']=int(input("Enter Base Salary"))

ws['B2']= "=B6*50%"

ws['B4']= "=B2*50%"

ws['B3'] = "=B6-B2-B4-B5"
ws['B5']="=B6*2.16%"
ws['C2']="=B2/12"
ws['C3']="=B3/12"
ws['C4']="=B4/12"
ws['C5']="=B5/12"
ws['B7']="=B6*5%"
ws['B8']="=B6*0.6%"
ws['C8']="=B8/12"
ws['B9']="=SUM(B6:B8)"
ws['C6']="=SUM(C2:C5)"
ws['C9']="=SUM(C6:C8)"   


wb.save("salary.xlsx")
